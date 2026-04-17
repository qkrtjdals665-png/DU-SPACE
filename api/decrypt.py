from http.server import BaseHTTPRequestHandler
import json, io, base64, hashlib

try:
    import msoffcrypto
    from openpyxl import load_workbook
except ImportError:
    msoffcrypto = None

class handler(BaseHTTPRequestHandler):
    def do_POST(self):
        try:
            content_length = int(self.headers.get('Content-Length', 0))
            body = self.rfile.read(content_length)
            req = json.loads(body)

            file_b64 = req.get('file', '')
            password = req.get('password', '900520')

            if not file_b64:
                return self._json(400, {'error': '파일이 없습니다'})

            if not msoffcrypto:
                return self._json(500, {'error': 'msoffcrypto 라이브러리가 설치되지 않았습니다'})

            # base64 → bytes
            file_bytes = base64.b64decode(file_b64)

            # 복호화 시도
            try:
                f = io.BytesIO(file_bytes)
                ms = msoffcrypto.OfficeFile(f)
                ms.load_key(password=password)
                decrypted = io.BytesIO()
                ms.decrypt(decrypted)
                decrypted.seek(0)
            except Exception:
                # 암호화 안 된 파일일 수 있음 - 그냥 통과
                decrypted = io.BytesIO(file_bytes)

            # openpyxl로 파싱
            wb = load_workbook(decrypted, data_only=True)
            result = {'sheets': []}

            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                rows = []
                for r in range(1, ws.max_row + 1):
                    row = []
                    for c in range(1, ws.max_column + 1):
                        val = ws.cell(r, c).value
                        if val is None:
                            row.append('')
                        elif hasattr(val, 'isoformat'):
                            row.append(val.isoformat())
                        else:
                            row.append(str(val))
                    rows.append(row)
                result['sheets'].append({
                    'name': sheet_name,
                    'rows': rows
                })

            return self._json(200, result)

        except Exception as e:
            return self._json(500, {'error': str(e)})

    def do_OPTIONS(self):
        self.send_response(200)
        self._cors_headers()
        self.end_headers()

    def _json(self, status, data):
        self.send_response(status)
        self._cors_headers()
        self.send_header('Content-Type', 'application/json; charset=utf-8')
        self.end_headers()
        self.wfile.write(json.dumps(data, ensure_ascii=False).encode('utf-8'))

    def _cors_headers(self):
        self.send_header('Access-Control-Allow-Origin', '*')
        self.send_header('Access-Control-Allow-Methods', 'POST, OPTIONS')
        self.send_header('Access-Control-Allow-Headers', 'Content-Type')
